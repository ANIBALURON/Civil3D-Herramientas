"""
+======================================================================+
|  MERGE CARTERA PREDOBLADO + COORDENADAS -> GIS                       |
|  TopografiaCivil3d.com                                               |
|                                                                      |
|  INSTALACION (una sola vez):                                         |
|    1. Abrir CMD como administrador                                   |
|    2. Escribir: pip install openpyxl                                  |
|    3. Listo                                                           |
|                                                                      |
|  USO:                                                                 |
|    Doble clic en este archivo .py                                     |
|    O desde CMD: python merge_cartera_gis.py                          |
+======================================================================+
"""

import os
import sys
import csv
import json
import math
import re
import configparser
from datetime import datetime

# ============================================================================
# VERIFICAR DEPENDENCIAS
# ============================================================================
try:
    import openpyxl
    TIENE_OPENPYXL = True
except ImportError:
    TIENE_OPENPYXL = False


# ============================================================================
# PROCESADOR DE DATOS
# ============================================================================
class CarteraGISProcessor:

    def __init__(self):
        self.csv_data = []
        self.excel_data = []
        self.merged_data = []
        self.log_messages = []
        self.tubos_repetidos = False
        self.config = None

    def log(self, msg):
        self.log_messages.append(msg)

    def get_and_clear_log(self):
        msgs = list(self.log_messages)
        self.log_messages.clear()
        return msgs

    def load_config(self, config_path):
        """Carga config_proyecto.ini y establece tubos_repetidos."""
        if not os.path.exists(config_path):
            self.log(f"  Config no encontrado: {config_path}")
            return False
        self.config = configparser.ConfigParser()
        self.config.read(config_path, encoding='utf-8')
        tr = self.config.get('PROYECTO', 'tubos_repetidos', fallback='no').strip().lower()
        self.tubos_repetidos = tr in ('si', 'yes', 'true', '1')
        self.log(f"  Config cargado: tubos_repetidos = {'SI' if self.tubos_repetidos else 'NO'}")
        return True

    def read_csv(self, filepath, separator=';'):
        """Lee un CSV de coordenadas. Retorna lista de registros."""
        data = []
        self.log(f"Leyendo CSV: {os.path.basename(filepath)}")

        with open(filepath, 'r', encoding='utf-8-sig') as f:
            first_line = f.readline()
            if ';' in first_line:
                separator = ';'
            elif ',' in first_line:
                separator = ','
            elif '\t' in first_line:
                separator = '\t'

        with open(filepath, 'r', encoding='utf-8-sig') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                parts = line.split(separator)
                if len(parts) < 5:
                    continue
                try:
                    punto_id = parts[0].strip()
                    este = float(parts[1].strip())
                    norte = float(parts[2].strip())
                    cota = float(parts[3].strip())
                    col5 = parts[4].strip()
                    col5_parts = col5.split()
                    num_tubo = col5_parts[0]
                    estacion_csv = col5_parts[1] if len(col5_parts) > 1 else ""
                    data.append({
                        'PuntoID': punto_id, 'Este': este, 'Norte': norte,
                        'Cota': cota, 'NumTubo': num_tubo, 'Estacion_CSV': estacion_csv
                    })
                except (ValueError, IndexError):
                    pass

        self.log(f"  -> {len(data)} puntos leidos")
        # Si no es modo tubos_repetidos, guardar en self.csv_data directamente
        if not self.tubos_repetidos:
            self.csv_data = data
        return data

    def read_csv_multiple(self, csv_folder, sheet_names):
        """Lee multiples CSV, uno por cada hoja. Retorna dict {nombre_hoja: [registros]}."""
        csv_by_sheet = {}
        files_in_folder = os.listdir(csv_folder)
        for sheet_name in sheet_names:
            found = False
            for fname in files_in_folder:
                if not fname.lower().endswith('.csv'):
                    continue
                if sheet_name.upper() in fname.upper():
                    csv_path = os.path.join(csv_folder, fname)
                    data = self.read_csv(csv_path)
                    csv_by_sheet[sheet_name] = data
                    found = True
                    break
            if not found:
                self.log(f"  AVISO: No se encontro CSV para hoja '{sheet_name}'")
                csv_by_sheet[sheet_name] = []
        return csv_by_sheet

    def read_excel(self, filepath):
        """Lee TODAS las hojas del Excel. Solo procesa hojas donde A1 contiene 'N. TUBO' o 'N.TUBO'."""
        self.excel_data = []
        self.log(f"Leyendo Excel: {os.path.basename(filepath)}")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True, keep_vba=False)

        def clean(val):
            if val is None:
                return ''
            s = str(val).strip()
            return '' if s == 'None' else s

        def clean_num(val):
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        hojas_datos = []
        hojas_ignoradas = []
        total_tubos = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Verificar si A1 contiene "N. TUBO" o "N.TUBO"
            first_cell = None
            for row in ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True):
                first_cell = str(row[0]).strip().upper() if row[0] is not None else ''
                break

            if first_cell and ('N. TUBO' in first_cell or 'N.TUBO' in first_cell):
                hojas_datos.append(sheet_name)
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    num_tubo = row[0] if len(row) > 0 else None
                    if num_tubo is None:
                        continue
                    num_tubo_str = str(num_tubo).strip()
                    if not num_tubo_str or num_tubo_str == 'None':
                        continue
                    # Si es float tipo 1.0, convertir a "1"
                    try:
                        num_float = float(num_tubo_str)
                        if num_float == int(num_float):
                            num_tubo_str = str(int(num_float))
                    except (ValueError, TypeError):
                        pass

                    self.excel_data.append({
                        'NumTubo': num_tubo_str,
                        'Longitud': clean_num(row[1] if len(row) > 1 else None),
                        'Serial': clean(row[2] if len(row) > 2 else None),
                        'Espesor': clean_num(row[3] if len(row) > 3 else None),
                        'GradosDir': clean(row[4] if len(row) > 4 else None),
                        'CentroCurva': clean(row[5] if len(row) > 5 else None),
                        'Notas': clean(row[6] if len(row) > 6 else None),
                        'PK_Junta': clean(row[7] if len(row) > 7 else None),
                        'Tramo': sheet_name,
                    })
                    count += 1
                total_tubos += count
                self.log(f"  Hoja '{sheet_name}': {count} tubos")
            else:
                hojas_ignoradas.append(sheet_name)

        wb.close()

        if hojas_ignoradas:
            self.log(f"  Hojas ignoradas: {', '.join(hojas_ignoradas)}")
        self.log(f"  -> TOTAL: {total_tubos} tubos de {len(hojas_datos)} hoja(s)")

        return total_tubos

    def merge_data(self, csv_folder=None):
        """Combina datos del Excel con coordenadas del CSV.

        Si tubos_repetidos=no: cruza por NumTubo contra self.csv_data (un solo CSV).
        Si tubos_repetidos=si: busca un CSV por hoja en csv_folder.
        """
        self.merged_data = []

        if self.tubos_repetidos and csv_folder:
            sheet_names = list(dict.fromkeys(r['Tramo'] for r in self.excel_data))
            csv_by_sheet = self.read_csv_multiple(csv_folder, sheet_names)
            for m in self.get_and_clear_log():
                self.log(m)

            matched = 0
            unmatched = 0
            for xr in self.excel_data:
                tramo = xr['Tramo']
                sheet_csv = csv_by_sheet.get(tramo, [])
                coords_dict = {}
                for r in sheet_csv:
                    t = r['NumTubo']
                    if t not in coords_dict:
                        coords_dict[t] = []
                    coords_dict[t].append(r)

                t = xr['NumTubo']
                if t in coords_dict:
                    for c in coords_dict[t]:
                        pk = c.get('Estacion_CSV') or xr.get('PK_Junta') or ''
                        self.merged_data.append({
                            'NumTubo': t, 'Este': c['Este'], 'Norte': c['Norte'],
                            'Cota': c['Cota'], 'Longitud': xr['Longitud'],
                            'Serial': xr['Serial'], 'Espesor': xr['Espesor'],
                            'GradosDir': xr['GradosDir'], 'CentroCurva': xr['CentroCurva'],
                            'Notas': xr['Notas'], 'PK_Junta': pk,
                            'Tramo': tramo,
                        })
                        matched += 1
                else:
                    unmatched += 1
        else:
            coords_dict = {}
            for r in self.csv_data:
                t = r['NumTubo']
                if t not in coords_dict:
                    coords_dict[t] = []
                coords_dict[t].append(r)

            # Detectar tubos duplicados en CSV
            dup_tubes = {t for t, v in coords_dict.items() if len(v) > 1}

            # PASO 1: asignar tubos unicos (sin ambiguedad)
            matched = 0
            unmatched = 0
            tramo_centroid = {}  # tramo -> lista de (este, norte)
            pending = []  # (xr, [coords]) para tubos duplicados

            for xr in self.excel_data:
                t = xr['NumTubo']
                if t not in coords_dict:
                    unmatched += 1
                    continue

                if t not in dup_tubes:
                    # Tubo unico en CSV: asignar directo
                    for c in coords_dict[t]:
                        pk = c.get('Estacion_CSV') or xr.get('PK_Junta') or ''
                        self.merged_data.append({
                            'NumTubo': t, 'Este': c['Este'], 'Norte': c['Norte'],
                            'Cota': c['Cota'], 'Longitud': xr['Longitud'],
                            'Serial': xr['Serial'], 'Espesor': xr['Espesor'],
                            'GradosDir': xr['GradosDir'], 'CentroCurva': xr['CentroCurva'],
                            'Notas': xr['Notas'], 'PK_Junta': pk,
                            'Tramo': xr.get('Tramo', ''),
                        })
                        matched += 1
                        tramo = xr.get('Tramo', '')
                        if tramo not in tramo_centroid:
                            tramo_centroid[tramo] = []
                        tramo_centroid[tramo].append((c['Este'], c['Norte']))
                else:
                    # Tubo duplicado: dejar para paso 2
                    pending.append((xr, coords_dict[t]))

            # PASO 2: tubos duplicados -> asignar la coordenada mas cercana al tramo
            matched_p2 = 0
            for xr, entries in pending:
                tramo = xr.get('Tramo', '')
                pts = tramo_centroid.get(tramo)
                if pts:
                    cx = sum(e for e, n in pts) / len(pts)
                    cy = sum(n for e, n in pts) / len(pts)
                    best = min(entries, key=lambda c: (c['Este'] - cx)**2 + (c['Norte'] - cy)**2)
                    pk = best.get('Estacion_CSV') or xr.get('PK_Junta') or ''
                    self.merged_data.append({
                        'NumTubo': xr['NumTubo'], 'Este': best['Este'], 'Norte': best['Norte'],
                        'Cota': best['Cota'], 'Longitud': xr['Longitud'],
                        'Serial': xr['Serial'], 'Espesor': xr['Espesor'],
                        'GradosDir': xr['GradosDir'], 'CentroCurva': xr['CentroCurva'],
                        'Notas': xr['Notas'], 'PK_Junta': pk,
                        'Tramo': tramo,
                    })
                    matched_p2 += 1
                else:
                    unmatched += 1

            matched += matched_p2
            if matched_p2:
                self.log(f"  Tubos duplicados resueltos por proximidad: {matched_p2}")

        self.log(f"\n  RESULTADO:")
        self.log(f"  Tubos combinados: {matched}")
        self.log(f"  Sin coordenadas: {unmatched}")
        return matched

    def export_geojson(self, output_path, epsg_code):
        features = []
        for row in self.merged_data:
            features.append({
                'type': 'Feature',
                'geometry': {'type': 'Point', 'coordinates': [row['Este'], row['Norte'], row['Cota']]},
                'properties': {
                    'N. TUBO': row['NumTubo'],
                    'LONGITUD': row['Longitud'],
                    'SERIAL NUMERO': row['Serial'],
                    'ESP. (mm)': row['Espesor'],
                    'GRADOS Y DIRECCION': row['GradosDir'],
                    'CENTRO DE LA CURVA': row['CentroCurva'],
                    'NOTAS': row['Notas'],
                    'PK/JUNTA': row['PK_Junta'],
                    'TRAMO': row.get('Tramo', ''),
                    'ESTE': row['Este'],
                    'NORTE': row['Norte'],
                    'COTA': row['Cota']
                }
            })

        geojson = {
            'type': 'FeatureCollection',
            'name': 'Cartera_Predoblado',
            'crs': {'type': 'name', 'properties': {'name': f'urn:ogc:def:crs:EPSG::{epsg_code}'}},
            'features': features
        }

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(geojson, f, indent=2, ensure_ascii=False)

        self.log(f"  GeoJSON: {os.path.basename(output_path)}")
        self.log(f"  CRS: EPSG:{epsg_code} | Tubos: {len(features)}")

    def export_kml(self, output_path, utm_zone):
        def utm_to_latlon(easting, northing, zone):
            a = 6378137.0; f = 1/298.257223563
            e = math.sqrt(2*f - f*f); e2 = e*e/(1-e*e); k0 = 0.9996
            x = easting - 500000; y = northing
            M = y/k0; mu = M/(a*(1-e*e/4-3*e**4/64-5*e**6/256))
            e1 = (1-math.sqrt(1-e*e))/(1+math.sqrt(1-e*e))
            phi1 = mu+(3*e1/2-27*e1**3/32)*math.sin(2*mu)
            phi1 += (21*e1**2/16-55*e1**4/32)*math.sin(4*mu)
            phi1 += (151*e1**3/96)*math.sin(6*mu)
            phi1 += (1097*e1**4/512)*math.sin(8*mu)
            N1 = a/math.sqrt(1-e*e*math.sin(phi1)**2)
            T1 = math.tan(phi1)**2; C1 = e2*math.cos(phi1)**2
            R1 = a*(1-e*e)/(1-e*e*math.sin(phi1)**2)**1.5
            D = x/(N1*k0)
            lat = phi1-(N1*math.tan(phi1)/R1)*(D**2/2-(5+3*T1+10*C1-4*C1**2-9*e2)*D**4/24+(61+90*T1+298*C1+45*T1**2-252*e2-3*C1**2)*D**6/720)
            lon = (D-(1+2*T1+C1)*D**3/6+(5-2*C1+28*T1-3*C1**2+8*e2+24*T1**2)*D**5/120)/math.cos(phi1)
            return math.degrees(lat), math.degrees(lon)+(zone*6-183)

        lines = ['<?xml version="1.0" encoding="UTF-8"?>',
                 '<kml xmlns="http://www.opengis.net/kml/2.2">','<Document>',
                 f'<name>Cartera Predoblado - {datetime.now().strftime("%Y-%m-%d")}</name>']
        for name, color, scale, icon in [('normal','ff00ff00','0.5','placemark_circle'),
                                          ('curva','ff0000ff','0.7','placemark_circle'),
                                          ('especial','ff00ffff','0.8','placemark_square')]:
            lines.append(f'<Style id="{name}"><IconStyle><color>{color}</color><scale>{scale}</scale><Icon><href>http://maps.google.com/mapfiles/kml/shapes/{icon}.png</href></Icon></IconStyle><LabelStyle><scale>0.7</scale></LabelStyle></Style>')

        for row in self.merged_data:
            lat, lon = utm_to_latlon(row['Este'], row['Norte'], utm_zone)
            curva = row.get('GradosDir',''); notas = row.get('Notas','')
            tramo = row.get('Tramo', '')
            style = '#especial' if any(k in notas.upper() for k in ['INICIO','FIN','PERFORADO','CODO']) else '#curva' if curva else '#normal'
            nm = f"T-{row['NumTubo']}" + (f" {curva}" if curva else "")
            desc = f'<![CDATA[<table border="1" cellpadding="3" style="font-size:11px;">'
            desc += f'<tr><td><b>Tubo</b></td><td>{row["NumTubo"]}</td></tr>'
            if tramo:
                desc += f'<tr><td><b>Tramo</b></td><td>{tramo}</td></tr>'
            if row.get('PK_Junta'): desc += f'<tr><td><b>PK</b></td><td>{row["PK_Junta"]}</td></tr>'
            if row.get('Longitud'): desc += f'<tr><td><b>Long</b></td><td>{row["Longitud"]}</td></tr>'
            if row.get('Serial'): desc += f'<tr><td><b>Serial</b></td><td>{row["Serial"]}</td></tr>'
            if curva: desc += f'<tr><td><b>Curva</b></td><td style="color:red;">{curva}</td></tr>'
            if row.get('CentroCurva'): desc += f'<tr><td><b>Centro</b></td><td>{row["CentroCurva"]}</td></tr>'
            if notas: desc += f'<tr><td><b>Notas</b></td><td style="color:blue;">{notas}</td></tr>'
            desc += f'<tr><td><b>Coord</b></td><td>{row["Este"]:.2f}, {row["Norte"]:.2f}, {row["Cota"]:.2f}</td></tr></table>]]>'
            lines.append(f'<Placemark><name>{nm}</name><description>{desc}</description><styleUrl>{style}</styleUrl><Point><coordinates>{lon:.8f},{lat:.8f},{row["Cota"]:.2f}</coordinates></Point></Placemark>')

        lines.extend(['</Document>','</kml>'])
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        self.log(f"  KML: {os.path.basename(output_path)} | Placemarks: {len(self.merged_data)}")

    def export_csv(self, output_path):
        headers = ['N. TUBO','LONGITUD','SERIAL NUMERO','ESP. (mm)','GRADOS Y DIRECCION',
                   'CENTRO DE LA CURVA','NOTAS','PK/JUNTA','TRAMO','ESTE','NORTE','COTA']
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(headers)
            for r in self.merged_data:
                w.writerow([r['NumTubo'], r['Longitud'], r['Serial'], r['Espesor'],
                            r['GradosDir'], r['CentroCurva'], r['Notas'], r['PK_Junta'],
                            r.get('Tramo', ''), r['Este'], r['Norte'], r['Cota']])
        self.log(f"  CSV: {os.path.basename(output_path)} | Registros: {len(self.merged_data)}")

    def export_by_tramo(self, output_folder, epsg_code):
        """Exporta un archivo por cada tramo del Excel.
        Tramos con coordenadas -> GeoJSON (capa espacial)
        Tramos sin coordenadas -> CSV (tabla sin geometria)
        Retorna dict {tramo: {'type': 'geojson'|'csv', 'path': path, 'count': int}}
        """
        os.makedirs(output_folder, exist_ok=True)

        # Todos los tramos del Excel, ordenados por PK (con coords primero, sin coords al final)
        all_tramos_set = list(dict.fromkeys(r['Tramo'] for r in self.excel_data if r.get('Tramo')))

        # Agrupar merged_data por tramo
        merged_by_tramo = {}
        for r in self.merged_data:
            tramo = r.get('Tramo', '')
            if tramo not in merged_by_tramo:
                merged_by_tramo[tramo] = []
            merged_by_tramo[tramo].append(r)

        # Agrupar excel_data por tramo (para tramos sin coords)
        excel_by_tramo = {}
        for r in self.excel_data:
            tramo = r.get('Tramo', '')
            if tramo not in excel_by_tramo:
                excel_by_tramo[tramo] = []
            excel_by_tramo[tramo].append(r)

        def sanitize(name):
            safe = name.replace(' ', '_')
            for c in '\\/:*?"<>|':
                safe = safe.replace(c, '_')
            return safe

        # Ordenar: con coordenadas por PK, sin coordenadas al final
        def _pk_sort(name):
            has_coords = 0 if name in merged_by_tramo and merged_by_tramo[name] else 1
            m_pk = re.search(r'KP\s*(\d+)\+(\d+)', name, re.IGNORECASE)
            pk_val = int(m_pk.group(1)) * 1000 + int(m_pk.group(2)) if m_pk else 999999
            return (has_coords, pk_val, name)

        all_tramos = sorted(all_tramos_set, key=_pk_sort)

        result = {}
        self.log(f"\n  EXPORTACION POR TRAMO:")

        for tramo in all_tramos:
            safe_name = sanitize(tramo)

            if tramo in merged_by_tramo and merged_by_tramo[tramo]:
                # Tiene coordenadas -> GeoJSON
                features = []
                for row in merged_by_tramo[tramo]:
                    features.append({
                        'type': 'Feature',
                        'geometry': {'type': 'Point', 'coordinates': [row['Este'], row['Norte'], row['Cota']]},
                        'properties': {
                            'N. TUBO': row['NumTubo'],
                            'LONGITUD': row['Longitud'],
                            'SERIAL NUMERO': row['Serial'],
                            'ESP. (mm)': row['Espesor'],
                            'GRADOS Y DIRECCION': row['GradosDir'],
                            'CENTRO DE LA CURVA': row['CentroCurva'],
                            'NOTAS': row['Notas'],
                            'PK/JUNTA': row['PK_Junta'],
                            'TRAMO': tramo,
                            'ESTE': row['Este'],
                            'NORTE': row['Norte'],
                            'COTA': row['Cota']
                        }
                    })
                geojson = {
                    'type': 'FeatureCollection',
                    'name': safe_name,
                    'crs': {'type': 'name', 'properties': {'name': f'urn:ogc:def:crs:EPSG::{epsg_code}'}},
                    'features': features
                }
                path = os.path.join(output_folder, f'cartera_{safe_name}.geojson')
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(geojson, f, indent=2, ensure_ascii=False)

                result[tramo] = {'type': 'geojson', 'path': path, 'count': len(features)}
                self.log(f"  {tramo}: {len(features)} tubos -> GeoJSON")
            else:
                # Sin coordenadas -> CSV (tabla)
                rows = excel_by_tramo.get(tramo, [])
                path = os.path.join(output_folder, f'tabla_{safe_name}.csv')
                headers = ['N. TUBO','LONGITUD','SERIAL NUMERO','ESP. (mm)',
                           'GRADOS Y DIRECCION','CENTRO DE LA CURVA','NOTAS','PK/JUNTA','TRAMO']
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    w = csv.writer(f, delimiter=';')
                    w.writerow(headers)
                    for r in rows:
                        w.writerow([r['NumTubo'], r['Longitud'], r['Serial'], r['Espesor'],
                                    r['GradosDir'], r['CentroCurva'], r['Notas'], r['PK_Junta'],
                                    tramo])

                result[tramo] = {'type': 'csv', 'path': path, 'count': len(rows)}
                self.log(f"  {tramo}: {len(rows)} tubos -> CSV (sin coordenadas)")

        self.log(f"  -> {len(result)} tramos exportados")
        return result

    def export_html_map(self, output_path, utm_zone):
        """Genera mapa HTML interactivo con Leaflet.js.
        Se abre en cualquier navegador sin necesidad de QGIS.
        """
        def utm_to_latlon(easting, northing, zone):
            a = 6378137.0; f = 1/298.257223563
            e = math.sqrt(2*f - f*f); e2 = e*e/(1-e*e); k0 = 0.9996
            x = easting - 500000; y = northing
            M = y/k0; mu = M/(a*(1-e*e/4-3*e**4/64-5*e**6/256))
            e1 = (1-math.sqrt(1-e*e))/(1+math.sqrt(1-e*e))
            phi1 = mu+(3*e1/2-27*e1**3/32)*math.sin(2*mu)
            phi1 += (21*e1**2/16-55*e1**4/32)*math.sin(4*mu)
            phi1 += (151*e1**3/96)*math.sin(6*mu)
            phi1 += (1097*e1**4/512)*math.sin(8*mu)
            N1 = a/math.sqrt(1-e*e*math.sin(phi1)**2)
            T1 = math.tan(phi1)**2; C1 = e2*math.cos(phi1)**2
            R1 = a*(1-e*e)/(1-e*e*math.sin(phi1)**2)**1.5
            D = x/(N1*k0)
            lat = phi1-(N1*math.tan(phi1)/R1)*(D**2/2-(5+3*T1+10*C1-4*C1**2-9*e2)*D**4/24+(61+90*T1+298*C1+45*T1**2-252*e2-3*C1**2)*D**6/720)
            lon = (D-(1+2*T1+C1)*D**3/6+(5-2*C1+28*T1-3*C1**2+8*e2+24*T1**2)*D**5/120)/math.cos(phi1)
            return math.degrees(lat), math.degrees(lon)+(zone*6-183)

        # Clasificar tipo de tubo y asignar color
        def get_tipo_color(row):
            notas = str(row.get('Notas', '') or '').upper()
            grados = str(row.get('GradosDir', '') or '').strip()
            centro = str(row.get('CentroCurva', '') or '').upper()
            pk_junta = str(row.get('PK_Junta', '') or '').upper()
            all_text = notas + ' ' + centro + ' ' + pk_junta + ' ' + grados.upper()
            if 'INICIO' in all_text:
                return 'INICIO', '#00cc00'
            if 'FIN' in all_text:
                return 'FIN', '#cc0000'
            if 'CODO' in notas:
                return 'Codo', '#8b0000'
            if 'NIPLE' in notas:
                return 'Niple', '#800080'
            if 'PERFORADO' in notas:
                return 'Perforado', '#ff8c00'
            if 'SOLAPE' in notas or 'TRASLAPE' in notas:
                return 'Solape', '#00cccc'
            if 'SECCION' in notas:
                return 'Seccion Movil', '#1E90FF'
            if 'CAMINO' in notas:
                return 'Camino', '#D2691E'
            if grados:
                return 'Curva', '#ffbf00'
            return 'Recto', '#999999'

        def parse_pk(pk_str):
            """Convierte '2+100' -> 2100.0, '62+839' -> 62839.0"""
            if not pk_str:
                return None
            pk = str(pk_str).strip().replace(',', '.')
            if '+' in pk:
                parts = pk.split('+')
                try:
                    return float(parts[0]) * 1000 + float(parts[1])
                except (ValueError, IndexError):
                    return None
            try:
                return float(pk)
            except ValueError:
                return None

        def haversine(lat1, lon1, lat2, lon2):
            """Distancia en metros entre dos puntos lat/lon."""
            R = 6371000
            phi1, phi2 = math.radians(lat1), math.radians(lat2)
            dphi = math.radians(lat2 - lat1)
            dlam = math.radians(lon2 - lon1)
            a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
            return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

        def format_pk(meters):
            """Formatea metros a formato PK: 2500 -> '2+500'"""
            km = int(meters // 1000)
            m = int(meters % 1000)
            return f"{km}+{m:03d}"

        # Funcion para ordenar tramos por PK del nombre
        def _tramo_sort_key(name):
            m_pk = re.search(r'KP\s*(\d+)\+(\d+)', name, re.IGNORECASE)
            return int(m_pk.group(1)) * 1000 + int(m_pk.group(2)) if m_pk else 999999

        # Convertir coordenadas y agrupar por tramo
        tramos_con_coords = {}
        all_lats = []
        all_lons = []
        for row in self.merged_data:
            lat, lon = utm_to_latlon(row['Este'], row['Norte'], utm_zone)
            all_lats.append(lat)
            all_lons.append(lon)
            tipo, color = get_tipo_color(row)
            tramo = row.get('Tramo', '') or 'Sin tramo'
            if tramo not in tramos_con_coords:
                tramos_con_coords[tramo] = []
            tramos_con_coords[tramo].append({
                'lat': lat, 'lon': lon,
                'tubo': row['NumTubo'],
                'pk': row.get('PK_Junta', ''),
                'serial': row.get('Serial', ''),
                'longitud': row.get('Longitud', ''),
                'grados': row.get('GradosDir', ''),
                'centro': row.get('CentroCurva', ''),
                'notas': row.get('Notas', ''),
                'espesor': row.get('Espesor', ''),
                'este': row['Este'],
                'norte': row['Norte'],
                'tramo': tramo,
                'tipo': tipo,
                'color': color,
            })

        # Tramos sin coordenadas (estan en excel_data pero no en merged_data)
        tramos_merged = set(r.get('Tramo', '') for r in self.merged_data)
        tramos_sin_coords = {}
        for r in self.excel_data:
            tramo = r.get('Tramo', '')
            if tramo and tramo not in tramos_merged:
                if tramo not in tramos_sin_coords:
                    tramos_sin_coords[tramo] = []
                tramos_sin_coords[tramo].append(r)

        # Ordenar tramos sin coordenadas por PK
        tramos_sin_coords = dict(sorted(tramos_sin_coords.items(), key=lambda x: _tramo_sort_key(x[0])))

        # Centro del mapa
        if all_lats:
            center_lat = sum(all_lats) / len(all_lats)
            center_lon = sum(all_lons) / len(all_lons)
        else:
            center_lat, center_lon = 19.4, -90.5

        # Ordenar tramos por PK
        tramos_con_coords = dict(sorted(tramos_con_coords.items(), key=lambda x: _tramo_sort_key(x[0])))

        # Generar datos JS por tramo
        tramos_js = []
        for tramo_name, puntos in tramos_con_coords.items():
            safe_name = tramo_name.replace("'", "\\'").replace('"', '\\"')
            markers_js = []
            for p in puntos:
                notas_safe = str(p['notas'] or '').replace("'", "\\'").replace('"', '&quot;').replace('\n', ' ')
                serial_safe = str(p['serial'] or '').replace("'", "\\'").replace('"', '&quot;')
                grados_safe = str(p['grados'] or '').replace("'", "\\'").replace('"', '&quot;')
                centro_safe = str(p['centro'] or '').replace("'", "\\'").replace('"', '&quot;')
                pk_safe = str(p['pk'] or '').replace("'", "\\'").replace('"', '&quot;')
                longitud_str = f"{p['longitud']:.2f}" if isinstance(p['longitud'], (int, float)) and p['longitud'] is not None else str(p['longitud'] or '')
                espesor_str = str(p['espesor'] or '')
                este_str = f"{p['este']:.3f}" if isinstance(p['este'], (int, float)) else str(p['este'] or '')
                norte_str = f"{p['norte']:.3f}" if isinstance(p['norte'], (int, float)) else str(p['norte'] or '')
                markers_js.append(
                    f"  m({p['lat']:.8f},{p['lon']:.8f},'{p['color']}','{p['tipo']}',"
                    f"'{p['tubo']}','{pk_safe}','{serial_safe}','{longitud_str}',"
                    f"'{grados_safe}','{centro_safe}','{notas_safe}','{safe_name}','{espesor_str}',"
                    f"'{este_str}','{norte_str}')"
                )
            tramos_js.append((safe_name, markers_js))

        # Generar resumen de tramos CON coordenadas
        con_coords_html = '<h3 style="margin:10px 0 5px;">Tramos con coordenadas</h3>'
        for tramo_name, puntos in tramos_con_coords.items():
            # Contar tipos
            tipos_count = {}
            for p in puntos:
                t = p['tipo']
                tipos_count[t] = tipos_count.get(t, 0) + 1
            detalle = ', '.join(f'{v} {k}' for k, v in tipos_count.items() if k not in ('Recto',))
            rectos = tipos_count.get('Recto', 0)
            if rectos:
                detalle = f'{rectos} rectos' + (f', {detalle}' if detalle else '')
            con_coords_html += f'<div style="font-size:11px;margin:2px 0;"><b style="color:#27ae60;">{tramo_name}</b>: {len(puntos)} tubos <span style="color:#888;">({detalle})</span></div>'

        # Generar tabla de tramos sin coordenadas
        sin_coords_html = ''
        if tramos_sin_coords:
            sin_coords_html = '<h3 style="margin:10px 0 5px;">Tramos sin coordenadas</h3>'
            for tramo_name, tubos in tramos_sin_coords.items():
                sin_coords_html += f'<details style="margin:5px 0;"><summary style="cursor:pointer;font-weight:bold;color:#c00;">{tramo_name} ({len(tubos)} tubos)</summary>'
                sin_coords_html += '<table class="sc-table"><tr><th>N.TUBO</th><th>Longitud</th><th>Serial</th><th>Grados</th><th>Notas</th></tr>'
                for t in tubos:
                    long_str = f"{t['Longitud']:.2f}" if isinstance(t['Longitud'], (int, float)) and t['Longitud'] is not None else ''
                    notas_esc = str(t.get('Notas', '') or '').replace('<', '&lt;').replace('>', '&gt;')
                    sin_coords_html += f"<tr><td>{t['NumTubo']}</td><td>{long_str}</td><td>{t.get('Serial','')}</td><td>{t.get('GradosDir','')}</td><td>{notas_esc}</td></tr>"
                sin_coords_html += '</table></details>'

        # Leyenda items
        leyenda_items = [
            ('INICIO', '#00cc00'), ('FIN', '#cc0000'), ('Codo', '#8b0000'),
            ('Niple', '#800080'), ('Perforado', '#ff8c00'), ('Solape', '#00cccc'),
            ('Secc. Movil', '#1E90FF'), ('Camino', '#D2691E'),
            ('Curva', '#ffbf00'), ('Recto', '#999999'),
        ]

        fecha = datetime.now().strftime('%Y-%m-%d %H:%M')

        # Escapar HTML para ponerlo dentro del JS
        con_coords_escaped = con_coords_html.replace('\\', '\\\\').replace('`', '\\`').replace('${', '\\${')
        sin_coords_escaped = sin_coords_html.replace('\\', '\\\\').replace('`', '\\`').replace('${', '\\${')

        # Descargar y cachear Leaflet para uso offline
        leaflet_head = ''
        cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.leaflet_cache')
        try:
            os.makedirs(cache_dir, exist_ok=True)
            js_path = os.path.join(cache_dir, 'leaflet.js')
            css_path = os.path.join(cache_dir, 'leaflet.css')
            if not os.path.exists(js_path) or not os.path.exists(css_path):
                import urllib.request
                self.log("  Descargando Leaflet.js para modo offline...")
                urllib.request.urlretrieve('https://unpkg.com/leaflet@1.9.4/dist/leaflet.js', js_path)
                urllib.request.urlretrieve('https://unpkg.com/leaflet@1.9.4/dist/leaflet.css', css_path)
                self.log("  Leaflet.js cacheado en .leaflet_cache/")
            with open(css_path, 'r', encoding='utf-8') as f:
                leaflet_css = f.read()
            with open(js_path, 'r', encoding='utf-8') as f:
                leaflet_js = f.read()
            leaflet_head = f'<style>{leaflet_css}</style>\n<script>{leaflet_js}</script>'
            self.log("  Leaflet embebido inline (funciona sin internet)")
        except Exception as ex:
            self.log(f"  No se pudo embeber Leaflet, usando CDN: {ex}")
            leaflet_head = '<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>\n<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>'

        html = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Mapa Cartera Predoblado - {fecha}</title>
{leaflet_head}
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; }}
  #map {{ position:absolute; top:0; left:0; right:0; bottom:0; z-index:1; }}
  .info-panel {{
    position:absolute; top:10px; right:10px; z-index:1000;
    background:rgba(255,255,255,0.95); border-radius:8px;
    box-shadow:0 2px 12px rgba(0,0,0,0.3); max-width:380px;
    max-height:calc(100vh - 20px); overflow-y:auto;
  }}
  .info-panel h2 {{ padding:12px 15px 8px; margin:0; font-size:14px; color:#333;
    border-bottom:1px solid #ddd; position:sticky; top:0; background:rgba(255,255,255,0.98); }}
  .info-panel .content {{ padding:8px 15px 12px; }}
  .legend {{ display:flex; flex-wrap:wrap; gap:6px 12px; margin:8px 0; }}
  .legend-item {{ display:flex; align-items:center; gap:4px; font-size:11px; }}
  .legend-dot {{ width:12px; height:12px; border-radius:50%; border:1px solid #555; flex-shrink:0; }}
  .stats {{ font-size:11px; color:#666; margin:5px 0; }}
  .toggle-btn {{
    background:#2196F3; color:#fff; border:none; border-radius:4px;
    padding:4px 10px; font-size:11px; cursor:pointer; margin:3px 2px;
  }}
  .toggle-btn:hover {{ background:#1976D2; }}
  .toggle-btn.active {{ background:#27ae60; }}
  .toggle-btn.active:hover {{ background:#219150; }}
  .label-ntubo {{ background:none !important; border:none !important; box-shadow:none !important; }}
  .label-ntubo span {{ font-size:10px; font-weight:bold; color:#fff; text-shadow:
    -1px -1px 0 #333, 1px -1px 0 #333, -1px 1px 0 #333, 1px 1px 0 #333; white-space:nowrap; }}
  .label-if {{ background:none !important; border:none !important; box-shadow:none !important; }}
  .label-if span {{ font-size:11px; font-weight:bold; padding:1px 4px; border-radius:3px; white-space:nowrap; }}
  .label-pk {{ background:none !important; border:none !important; box-shadow:none !important; }}
  .label-pk span {{ font-size:10px; font-weight:bold; color:#00FFFF; background:rgba(0,0,0,0.65);
    padding:1px 5px; border-radius:3px; white-space:nowrap; border:1px solid #00FFFF; }}
  .sc-table {{ font-size:10px; border-collapse:collapse; width:100%; margin:5px 0; }}
  .sc-table th, .sc-table td {{ border:1px solid #ccc; padding:2px 5px; text-align:left; }}
  .sc-table th {{ background:#f0f0f0; }}
  .leaflet-popup-content {{ margin:8px; }}
  .popup-table {{ border-collapse:collapse; font-size:11px; }}
  .popup-table td {{ border:1px solid #ddd; padding:3px 8px; }}
  .popup-table td:first-child {{ font-weight:bold; background:#f5f5f5; white-space:nowrap; }}
  .popup-table .highlight {{ color:#cc0000; font-weight:bold; }}
  .popup-btns {{ display:flex; gap:5px; margin-top:6px; }}
  .popup-btns a, .popup-btns button {{ flex:1; padding:6px 4px; border:none; border-radius:4px; font-size:11px; font-weight:bold; cursor:pointer; text-align:center; text-decoration:none; }}
  .btn-navegar {{ background:#4285F4; color:#fff !important; }}
  .btn-copiar {{ background:#34A853; color:#fff; }}
  .btn-copiar.copiado {{ background:#888; }}
  .leaflet-control-layers-toggle {{
    width:auto !important; height:auto !important;
    padding:6px 10px !important; font-size:12px; font-weight:bold;
    color:#333; background:#fff !important; background-image:none !important;
  }}
</style>
</head>
<body>
<div id="map"></div>
<div class="info-panel" id="panel">
  <h2>Cartera Predoblado
    <button class="toggle-btn" onclick="document.getElementById('panel').style.display='none'" style="float:right;background:#999;padding:2px 8px;">X</button>
  </h2>
  <div class="content">
    <div class="stats">{len(self.merged_data)} tubos con coordenadas | {sum(len(v) for v in tramos_sin_coords.values())} sin coordenadas</div>
    <div class="stats">Generado: {fecha}</div>
    <div class="legend">
      {"".join(f'<span class="legend-item"><span class="legend-dot" style="background:{c}"></span>{n}</span>' for n, c in leyenda_items)}
    </div>
    <h3 style="margin:10px 0 5px;">Etiquetas</h3>
    <div>
      <button class="toggle-btn" id="btnLblTubo" onclick="toggleLabels('tubo')">N. TUBO</button>
      <button class="toggle-btn" id="btnLblIF" onclick="toggleLabels('if')">ESPECIALES</button>
      <button class="toggle-btn" id="btnLblPK" onclick="toggleLabels('pk')">PK</button>
    </div>
    <h3 style="margin:10px 0 5px;">Modo campo (sin internet)</h3>
    <div>
      <button class="toggle-btn" id="btnOffline" onclick="downloadForField()" style="background:#E65100;">Guardar para campo</button>
      <button class="toggle-btn" onclick="clearTileCache()" style="background:#999;font-size:10px;">Borrar cache</button>
    </div>
    <div id="dlProgress" style="display:none;margin:5px 0;">
      <div style="background:#ddd;border-radius:4px;height:18px;overflow:hidden;">
        <div id="dlBar" style="background:#4CAF50;height:100%;width:0%;transition:width 0.2s;text-align:center;color:#fff;font-size:10px;line-height:18px;">0%</div>
      </div>
      <div id="dlStatus" style="font-size:10px;color:#666;margin-top:2px;"></div>
    </div>
    <div id="con-coords"></div>
    <div id="sin-coords"></div>
  </div>
</div>
<script>
// Mapa base
var map = L.map('map',{{zoomControl:true,preferCanvas:true,tapTolerance:20}}).setView([{center_lat:.6f},{center_lon:.6f}],13);
map.createPane('linePane');
map.getPane('linePane').style.zIndex = 350;

// --- IndexedDB para cache de tiles offline ---
var tileDB=null;
(function(){{
  var req=indexedDB.open('mapTilesCache',1);
  req.onupgradeneeded=function(e){{var db=e.target.result;if(!db.objectStoreNames.contains('tiles'))db.createObjectStore('tiles');}};
  req.onsuccess=function(e){{tileDB=e.target.result;}};
}})();
function saveTile(k,b){{if(!tileDB)return;try{{var t=tileDB.transaction('tiles','readwrite');t.objectStore('tiles').put(b,k);}}catch(e){{}}}}
function getTile(k,cb){{if(!tileDB){{cb(null);return;}}try{{var t=tileDB.transaction('tiles','readonly');var r=t.objectStore('tiles').get(k);r.onsuccess=function(){{cb(r.result||null);}};r.onerror=function(){{cb(null);}};}}catch(e){{cb(null);}}}}

// Tile layer con cache offline (para capas con CORS: ESRI, OSM)
L.TileLayer.Offline=L.TileLayer.extend({{
  createTile:function(coords,done){{
    var tile=document.createElement('img');
    var url=this.getTileUrl(coords);
    getTile(url,function(blob){{
      if(blob){{tile.src=URL.createObjectURL(blob);done(null,tile);}}
      else{{
        tile.crossOrigin='anonymous';
        tile.onload=function(){{
          try{{var c=document.createElement('canvas');c.width=tile.naturalWidth;c.height=tile.naturalHeight;
          c.getContext('2d').drawImage(tile,0,0);c.toBlob(function(b){{if(b)saveTile(url,b);}});}}catch(e){{}}
          done(null,tile);
        }};
        tile.onerror=function(){{done('error',tile);}};
        tile.src=url;
      }}
    }});
    return tile;
  }}
}});

// Capas base
var esri = new L.TileLayer.Offline('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}',{{
  attribution:'ESRI Satellite',maxZoom:18
}});
var gsat = L.tileLayer('https://mt1.google.com/vt/lyrs=y&x={{x}}&y={{y}}&z={{z}}',{{
  attribution:'Google Satellite (online)',maxZoom:20
}});
var osm = new L.TileLayer.Offline('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',{{
  attribution:'&copy; OpenStreetMap',maxZoom:19
}});
esri.addTo(map);

var baseMaps = {{"ESRI Satelite (offline)":esri, "Google Satelite (online)":gsat, "OpenStreetMap (offline)":osm}};
var overlays = {{}};
var allMarkers = [];
var lblTuboLayer = L.layerGroup();
var lblIFLayer = L.layerGroup();
var lblPKLayer = L.layerGroup();
var lblLineLayer = L.layerGroup();
var lblTuboOn = false;
var lblIFOn = false;
var lblPKOn = false;
var lblLineOn = false;

function toggleLabels(tipo) {{
  if(tipo==='tubo') {{
    lblTuboOn = !lblTuboOn;
    if(lblTuboOn) {{ lblTuboLayer.addTo(map); document.getElementById('btnLblTubo').classList.add('active'); }}
    else {{ map.removeLayer(lblTuboLayer); document.getElementById('btnLblTubo').classList.remove('active'); }}
  }} else if(tipo==='if') {{
    lblIFOn = !lblIFOn;
    if(lblIFOn) {{ lblIFLayer.addTo(map); document.getElementById('btnLblIF').classList.add('active'); }}
    else {{ map.removeLayer(lblIFLayer); document.getElementById('btnLblIF').classList.remove('active'); }}
  }} else if(tipo==='pk') {{
    lblPKOn = !lblPKOn;
    if(lblPKOn) {{ lblPKLayer.addTo(map); document.getElementById('btnLblPK').classList.add('active'); }}
    else {{ map.removeLayer(lblPKLayer); document.getElementById('btnLblPK').classList.remove('active'); }}
  }} else if(tipo==='line') {{
    lblLineOn = !lblLineOn;
    if(lblLineOn) {{ lblLineLayer.addTo(map); document.getElementById('btnLblLine').classList.add('active'); }}
    else {{ map.removeLayer(lblLineLayer); document.getElementById('btnLblLine').classList.remove('active'); }}
  }}
}}

// Copiar coordenadas al portapapeles
function copiarCoord(btn,lat,lon) {{
  var txt = lat.toFixed(6)+','+lon.toFixed(6);
  navigator.clipboard.writeText(txt).then(function(){{
    btn.textContent='Copiado!'; btn.classList.add('copiado');
    setTimeout(function(){{ btn.textContent='Copiar coord.'; btn.classList.remove('copiado'); }},2000);
  }});
}}

// Funcion para crear marcadores
function m(lat,lon,color,tipo,tubo,pk,serial,longitud,grados,centro,notas,tramo,espesor,este,norte) {{
  var circle = L.circleMarker([lat,lon],{{
    radius: (tipo==='INICIO'||tipo==='FIN') ? 10 : (tipo==='Curva'||tipo==='Codo') ? 9 : (tipo==='Recto' ? 6 : 8),
    fillColor: color,
    color: '#333',
    weight: 2,
    opacity: 1,
    fillOpacity: 0.85
  }});
  var html = '<table class="popup-table">';
  html += '<tr><td>N. TUBO</td><td><b>'+tubo+'</b></td></tr>';
  html += '<tr><td>Tramo</td><td>'+tramo+'</td></tr>';
  if(pk) html += '<tr><td>PK/Junta</td><td>'+pk+'</td></tr>';
  if(serial) html += '<tr><td>Serial</td><td>'+serial+'</td></tr>';
  if(longitud) html += '<tr><td>Longitud</td><td>'+longitud+' m</td></tr>';
  if(espesor) html += '<tr><td>Espesor</td><td>'+espesor+' mm</td></tr>';
  if(grados) html += '<tr><td>Grados</td><td class="highlight">'+grados+'</td></tr>';
  if(centro) html += '<tr><td>Centro curva</td><td>'+centro+'</td></tr>';
  if(notas) html += '<tr><td>Notas</td><td class="highlight">'+notas+'</td></tr>';
  if(tipo==='INICIO' && este) html += '<tr><td>Este</td><td>'+este+'</td></tr>';
  if(tipo==='INICIO' && norte) html += '<tr><td>Norte</td><td>'+norte+'</td></tr>';
  html += '<tr><td>Tipo</td><td><span style="color:'+color+';font-weight:bold;">'+tipo+'</span></td></tr>';
  html += '</table>';
  html += '<div class="popup-btns">';
  html += '<a class="btn-navegar" href="https://www.google.com/maps/dir/?api=1&destination='+lat+','+lon+'&travelmode=driving" target="_blank">Como llegar</a>';
  html += '<button class="btn-copiar" onclick="copiarCoord(this,'+lat+','+lon+')">Copiar coord.</button>';
  html += '</div>';
  circle.bindPopup(html,{{maxWidth:300}});
  allMarkers.push(circle);
  // Etiqueta N.TUBO
  var lblT = L.marker([lat,lon],{{icon:L.divIcon({{className:'label-ntubo',html:'<span>'+tubo+'</span>',iconSize:[0,0],iconAnchor:[-8,-8]}})
  }});
  lblTuboLayer.addLayer(lblT);
  // Etiqueta ESPECIALES (todo lo que no es Recto ni Curva)
  if(tipo!=='Recto' && tipo!=='Curva') {{
    var lblIF = L.marker([lat,lon],{{icon:L.divIcon({{className:'label-if',html:'<span style="background:'+color+';color:#fff;">T-'+tubo+' '+tipo+'</span>',iconSize:[0,0],iconAnchor:[-10,12]}})
    }});
    lblIFLayer.addLayer(lblIF);
  }}
  // Etiqueta PK
  if(pk) {{
    var lblPK = L.marker([lat,lon],{{icon:L.divIcon({{className:'label-pk',html:'<span>'+pk+'</span>',iconSize:[0,0],iconAnchor:[-8,28]}})
    }});
    lblPKLayer.addLayer(lblPK);
  }}
  return circle;
}}

// Crear capas por tramo
'''
        # Agregar marcadores por tramo
        for tramo_name, markers in tramos_js:
            var_name = 'tr_' + ''.join(c if c.isalnum() else '_' for c in tramo_name)
            html += f"var {var_name} = L.layerGroup([\n"
            html += ',\n'.join(markers)
            html += f"\n]).addTo(map);\noverlays['{tramo_name}'] = {var_name};\n\n"


        html += f'''
// Control de capas
var layerCtrl = L.control.layers(baseMaps, overlays, {{collapsed:true}}).addTo(map);
// Texto en el boton de capas
var toggle = layerCtrl.getContainer().querySelector('.leaflet-control-layers-toggle');
if(toggle) toggle.textContent = 'Capas / Tramos';

// Auto-zoom al extent
if(allMarkers.length > 0) {{
  var group = L.featureGroup(allMarkers);
  map.fitBounds(group.getBounds().pad(0.05));
}}

// Panel tramos
document.getElementById('con-coords').innerHTML = `{con_coords_escaped}`;
document.getElementById('sin-coords').innerHTML = `{sin_coords_escaped}`;

// Boton para mostrar panel (si se cierra)
var showBtn = L.control({{position:'topright'}});
showBtn.onAdd = function() {{
  var div = L.DomUtil.create('div');
  div.innerHTML = '<button onclick="document.getElementById(\\'panel\\').style.display=\\'block\\'" style="background:#fff;border:1px solid #ccc;border-radius:4px;padding:5px 10px;cursor:pointer;font-size:12px;box-shadow:0 1px 5px rgba(0,0,0,.3)">Info</button>';
  return div;
}};
showBtn.addTo(map);

// Escala
L.control.scale({{imperial:false}}).addTo(map);

// GPS - Mi ubicacion
var gpsMarker = null;
var gpsCircle = null;
var gpsWatchId = null;
var gpsBtn = L.control({{position:'topleft'}});
gpsBtn.onAdd = function() {{
  var div = L.DomUtil.create('div','leaflet-bar');
  div.innerHTML = '<a href="#" id="gpsBtn" title="Mi ubicacion" style="font-size:18px;line-height:30px;width:30px;height:30px;display:block;text-align:center;text-decoration:none;color:#333;background:#fff;">&#9737;</a>';
  div.onclick = function(e) {{
    e.preventDefault();
    e.stopPropagation();
    if(gpsWatchId !== null) {{
      navigator.geolocation.clearWatch(gpsWatchId);
      gpsWatchId = null;
      if(gpsMarker) map.removeLayer(gpsMarker);
      if(gpsCircle) map.removeLayer(gpsCircle);
      gpsMarker = null; gpsCircle = null;
      document.getElementById('gpsBtn').style.color = '#333';
      return;
    }}
    if(!navigator.geolocation) {{ alert('GPS no disponible'); return; }}
    document.getElementById('gpsBtn').style.color = '#2196F3';
    gpsWatchId = navigator.geolocation.watchPosition(function(pos) {{
      var lat = pos.coords.latitude;
      var lon = pos.coords.longitude;
      var acc = pos.coords.accuracy;
      if(!gpsMarker) {{
        gpsMarker = L.circleMarker([lat,lon],{{radius:8,fillColor:'#2196F3',color:'#fff',weight:3,fillOpacity:1}}).addTo(map);
        gpsCircle = L.circle([lat,lon],{{radius:acc,color:'#2196F3',fillColor:'#2196F3',fillOpacity:0.1,weight:1}}).addTo(map);
        map.setView([lat,lon],17);
      }} else {{
        gpsMarker.setLatLng([lat,lon]);
        gpsCircle.setLatLng([lat,lon]);
        gpsCircle.setRadius(acc);
      }}
      gpsMarker.bindPopup('Mi ubicacion<br>Precision: '+Math.round(acc)+' m');
    }}, function(err) {{
      alert('Error GPS: '+err.message);
      document.getElementById('gpsBtn').style.color = '#c00';
    }}, {{enableHighAccuracy:true, maximumAge:5000, timeout:10000}});
  }};
  return div;
}};
gpsBtn.addTo(map);

// --- Descargar tiles para campo (offline) ---
function latLngToTile(lat,lng,z){{
  var n=Math.pow(2,z);
  var x=Math.floor((lng+180)/360*n);
  var y=Math.floor((1-Math.log(Math.tan(lat*Math.PI/180)+1/Math.cos(lat*Math.PI/180))/Math.PI)/2*n);
  return {{x:x,y:y}};
}}

function downloadForField(){{
  var bounds=map.getBounds();
  var curZ=map.getZoom();
  var minZ=Math.max(10,curZ-2);
  var maxZ=Math.min(17,curZ+3);

  // Calcular tiles a descargar (ESRI + OSM)
  var tiles=[];
  for(var z=minZ;z<=maxZ;z++){{
    var tl=latLngToTile(bounds.getNorth(),bounds.getWest(),z);
    var br=latLngToTile(bounds.getSouth(),bounds.getEast(),z);
    for(var x=tl.x;x<=br.x;x++){{
      for(var y=tl.y;y<=br.y;y++){{
        tiles.push('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/'+z+'/'+y+'/'+x);
        var s=['a','b','c'][Math.abs(x+y)%3];
        tiles.push('https://'+s+'.tile.openstreetmap.org/'+z+'/'+x+'/'+y+'.png');
      }}
    }}
  }}

  if(tiles.length>5000){{
    if(!confirm('Se van a descargar '+tiles.length+' tiles (puede tardar). ¿Continuar?'))return;
  }}

  var prog=document.getElementById('dlProgress');
  var bar=document.getElementById('dlBar');
  var status=document.getElementById('dlStatus');
  var btn=document.getElementById('btnOffline');
  prog.style.display='block';
  btn.disabled=true;
  btn.textContent='Descargando...';

  var done=0,errors=0,total=tiles.length;
  var batch=6; // descargas simultaneas
  var idx=0;

  function updateUI(){{
    var pct=Math.round(done/total*100);
    bar.style.width=pct+'%';
    bar.textContent=pct+'%';
    status.textContent=done+'/'+total+' tiles ('+errors+' errores)';
  }}

  function next(){{
    if(idx>=total){{
      if(done+errors>=total){{
        btn.disabled=false;
        btn.textContent='Guardar para campo';
        btn.style.background='#27ae60';
        status.textContent='Listo: '+(done-errors)+' tiles guardados. Puede usar el mapa sin internet.';
      }}
      return;
    }}
    var url=tiles[idx++];
    var img=new Image();
    img.crossOrigin='anonymous';
    img.onload=function(){{
      try{{
        var c=document.createElement('canvas');c.width=img.naturalWidth;c.height=img.naturalHeight;
        c.getContext('2d').drawImage(img,0,0);
        c.toBlob(function(b){{if(b)saveTile(url,b);done++;updateUI();next();}});
      }}catch(e){{done++;errors++;updateUI();next();}}
    }};
    img.onerror=function(){{done++;errors++;updateUI();next();}};
    img.src=url;
  }}

  // Iniciar descargas en paralelo
  for(var i=0;i<batch&&i<total;i++)next();
}}

function clearTileCache(){{
  if(!confirm('¿Borrar todos los tiles guardados?'))return;
  indexedDB.deleteDatabase('mapTilesCache');
  tileDB=null;
  var req=indexedDB.open('mapTilesCache',1);
  req.onupgradeneeded=function(e){{e.target.result.createObjectStore('tiles');}};
  req.onsuccess=function(e){{tileDB=e.target.result;}};
  document.getElementById('dlStatus').textContent='Cache borrado.';
  document.getElementById('btnOffline').style.background='#E65100';
}}
</script>
</body>
</html>'''

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

        self.log(f"  Mapa HTML: {os.path.basename(output_path)}")
        self.log(f"  Tubos en mapa: {len(self.merged_data)} | Tramos: {len(tramos_con_coords)}")
        if tramos_sin_coords:
            self.log(f"  Sin coordenadas: {sum(len(v) for v in tramos_sin_coords.values())} tubos en {len(tramos_sin_coords)} tramo(s)")


# ============================================================================
# INTERFAZ GRAFICA - solo se ejecuta en modo standalone (doble clic)
# ============================================================================
if __name__ == '__main__':
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Merge Cartera Predoblado -> GIS  |  TopografiaCivil3d.com")
            self.geometry("780x720")
            self.configure(bg='#1e1e2e')
            self.resizable(False, False)
            self.processor = CarteraGISProcessor()
            self.csv_path = tk.StringVar()
            self.excel_path = tk.StringVar()
            self.output_format = tk.StringVar(value="geojson")
            self.utm_zone = tk.IntVar(value=16)
            self.tubos_rep_var = tk.BooleanVar(value=False)
            self._build_ui()
            self._try_load_config()

        def _try_load_config(self):
            script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in dir() else os.getcwd()
            config_path = os.path.join(script_dir, 'config_proyecto.ini')
            if os.path.exists(config_path):
                self.processor.load_config(config_path)
                self.tubos_rep_var.set(self.processor.tubos_repetidos)
                for m in self.processor.get_and_clear_log():
                    self._log(m)
                cfg = self.processor.config
                if cfg:
                    project_root = os.path.dirname(script_dir)
                    csv_rel = cfg.get('ARCHIVOS', 'coordenadas_csv', fallback='')
                    xl_rel = cfg.get('ARCHIVOS', 'cartera_excel', fallback='')
                    if csv_rel:
                        csv_abs = os.path.join(project_root, csv_rel)
                        if os.path.exists(csv_abs):
                            self.csv_path.set(csv_abs)
                            self._log(f"  CSV auto: {os.path.basename(csv_abs)}")
                    if xl_rel:
                        xl_abs = os.path.join(project_root, xl_rel)
                        if os.path.exists(xl_abs):
                            self.excel_path.set(xl_abs)
                            self._log(f"  Excel auto: {os.path.basename(xl_abs)}")
                    zone = cfg.getint('PROYECTO', 'zona_utm', fallback=16)
                    self.utm_zone.set(zone)

        def _build_ui(self):
            BG='#1e1e2e'; FG='#cdd6f4'; ACCENT='#89b4fa'; BTN_BG='#313244'
            ENTRY_BG='#181825'; GREEN='#a6e3a1'; YELLOW='#f9e2af'

            tk.Label(self, text="MERGE CARTERA PREDOBLADO -> GIS",
                     font=('Consolas',15,'bold'), fg=ACCENT, bg=BG).pack(anchor='w', padx=20, pady=(15,0))
            tk.Label(self, text="Combina cartera (Excel) + coordenadas (CSV) y exporta para QGIS / ArcGIS / Google Earth",
                     font=('Segoe UI',9), fg='#6c7086', bg=BG).pack(anchor='w', padx=20, pady=(0,10))

            f1 = tk.LabelFrame(self, text=" 1. ARCHIVOS DE ENTRADA ", font=('Segoe UI',10,'bold'), fg=ACCENT, bg=BG, bd=1)
            f1.pack(fill='x', padx=20, pady=5)

            row1 = tk.Frame(f1, bg=BG); row1.pack(fill='x', padx=10, pady=5)
            tk.Label(row1, text="CSV Coordenadas:", font=('Segoe UI',9,'bold'), fg=YELLOW, bg=BG, width=18, anchor='w').pack(side='left')
            tk.Entry(row1, textvariable=self.csv_path, font=('Consolas',9), bg=ENTRY_BG, fg=YELLOW, insertbackground=YELLOW, bd=0, highlightthickness=1, highlightcolor=ACCENT).pack(side='left', fill='x', expand=True, padx=5)
            tk.Button(row1, text="Buscar", command=self._browse_csv, bg=BTN_BG, fg=FG, bd=0, padx=12, cursor='hand2').pack(side='left')
            tk.Label(f1, text="    Formato: PuntoID ; Este ; Norte ; Cota ; NumTubo [Estacion]", font=('Consolas',8), fg='#585b70', bg=BG).pack(anchor='w', padx=10)

            row2 = tk.Frame(f1, bg=BG); row2.pack(fill='x', padx=10, pady=5)
            tk.Label(row2, text="Excel Cartera:", font=('Segoe UI',9,'bold'), fg=YELLOW, bg=BG, width=18, anchor='w').pack(side='left')
            tk.Entry(row2, textvariable=self.excel_path, font=('Consolas',9), bg=ENTRY_BG, fg=YELLOW, insertbackground=YELLOW, bd=0, highlightthickness=1, highlightcolor=ACCENT).pack(side='left', fill='x', expand=True, padx=5)
            tk.Button(row2, text="Buscar", command=self._browse_excel, bg=BTN_BG, fg=FG, bd=0, padx=12, cursor='hand2').pack(side='left')
            tk.Label(f1, text="    Columnas: N.Tubo | Longitud | Serial | Espesor | Grados | Centro | Notas | PK", font=('Consolas',8), fg='#585b70', bg=BG).pack(anchor='w', padx=10, pady=(0,5))

            f2 = tk.LabelFrame(self, text=" 2. OPCIONES ", font=('Segoe UI',10,'bold'), fg=ACCENT, bg=BG, bd=1)
            f2.pack(fill='x', padx=20, pady=5)

            utm_row = tk.Frame(f2, bg=BG); utm_row.pack(fill='x', padx=10, pady=5)
            tk.Label(utm_row, text="Zona UTM:", font=('Segoe UI',9,'bold'), fg=FG, bg=BG).pack(side='left', padx=(0,15))
            tk.Radiobutton(utm_row, text="Zona 16N (Yucatan / Campeche / Q.Roo)", variable=self.utm_zone, value=16, font=('Segoe UI',9), fg=GREEN, bg=BG, selectcolor=ENTRY_BG, activebackground=BG).pack(side='left', padx=5)
            tk.Radiobutton(utm_row, text="Zona 15N (Tabasco / Chiapas / Oaxaca)", variable=self.utm_zone, value=15, font=('Segoe UI',9), fg=GREEN, bg=BG, selectcolor=ENTRY_BG, activebackground=BG).pack(side='left', padx=5)

            fmt_row = tk.Frame(f2, bg=BG); fmt_row.pack(fill='x', padx=10, pady=5)
            tk.Label(fmt_row, text="Formato:", font=('Segoe UI',9,'bold'), fg=FG, bg=BG).pack(side='left', padx=(0,10))
            for text, val in [("GeoJSON (QGIS/ArcGIS)","geojson"),("KML (Google Earth)","kml"),("CSV combinado","csv"),("Todos","all")]:
                tk.Radiobutton(fmt_row, text=text, variable=self.output_format, value=val, font=('Segoe UI',9), fg=FG, bg=BG, selectcolor=ENTRY_BG, activebackground=BG).pack(side='left', padx=5)

            rep_row = tk.Frame(f2, bg=BG); rep_row.pack(fill='x', padx=10, pady=(0,5))
            tk.Label(rep_row, text="Modo:", font=('Segoe UI',9,'bold'), fg=FG, bg=BG).pack(side='left', padx=(0,10))
            self.rep_check = tk.Checkbutton(rep_row, text="Tubos repetidos (un CSV por hoja)",
                                             variable=self.tubos_rep_var, font=('Segoe UI',9),
                                             fg='#fab387', bg=BG, selectcolor=ENTRY_BG, activebackground=BG)
            self.rep_check.pack(side='left', padx=5)

            tk.Button(self, text="PROCESAR Y EXPORTAR", command=self._process, font=('Segoe UI',13,'bold'), bg='#89b4fa', fg='#1e1e2e', activebackground='#b4d0fb', bd=0, padx=20, pady=10, cursor='hand2').pack(fill='x', padx=20, pady=10)

            f3 = tk.LabelFrame(self, text=" LOG ", font=('Segoe UI',10,'bold'), fg=ACCENT, bg=BG, bd=1)
            f3.pack(fill='both', expand=True, padx=20, pady=(5,15))
            self.log_text = tk.Text(f3, font=('Consolas',9), bg=ENTRY_BG, fg=GREEN, insertbackground=GREEN, bd=0, wrap='word')
            sb = tk.Scrollbar(f3, command=self.log_text.yview)
            self.log_text.configure(yscrollcommand=sb.set)
            sb.pack(side='right', fill='y')
            self.log_text.pack(fill='both', expand=True, padx=5, pady=5)

            self._log("=== Merge Cartera Predoblado -> GIS ===")
            self._log("TopografiaCivil3d.com\n")
            if TIENE_OPENPYXL:
                self._log("OK openpyxl instalado\n")
            else:
                self._log("X openpyxl NO instalado")
                self._log("  Abra CMD: pip install openpyxl\n")
            self._log("1. Seleccione CSV de coordenadas")
            self._log("2. Seleccione Excel de cartera")
            self._log("3. Elija zona UTM (15 o 16) y formato")
            self._log("4. Presione PROCESAR")
            self._log("5. Abra el archivo .geojson en QGIS\n")

        def _log(self, msg):
            self.log_text.insert('end', msg + '\n')
            self.log_text.see('end')
            self.update_idletasks()

        def _browse_csv(self):
            p = filedialog.askopenfilename(title="CSV de Coordenadas", filetypes=[("CSV","*.csv"),("Texto","*.txt"),("Todos","*.*")])
            if p: self.csv_path.set(p); self._log(f"CSV: {os.path.basename(p)}")

        def _browse_excel(self):
            p = filedialog.askopenfilename(title="Excel de Cartera", filetypes=[("Excel","*.xlsx *.xlsm *.xls"),("Todos","*.*")])
            if p: self.excel_path.set(p); self._log(f"Excel: {os.path.basename(p)}")

        def _process(self):
            csv_f = self.csv_path.get(); xl_f = self.excel_path.get()
            if not csv_f or not os.path.exists(csv_f):
                messagebox.showerror("Error","Seleccione un CSV valido."); return
            if not xl_f or not os.path.exists(xl_f):
                messagebox.showerror("Error","Seleccione un Excel valido."); return
            if not TIENE_OPENPYXL:
                messagebox.showerror("Error","Instale openpyxl:\npip install openpyxl"); return

            self._log("\n" + "=" * 50)
            self._log("PROCESANDO...")
            self._log("=" * 50)

            self.processor = CarteraGISProcessor()
            self.processor.tubos_repetidos = self.tubos_rep_var.get()

            try:
                if not self.processor.tubos_repetidos:
                    self.processor.read_csv(csv_f)
                    for m in self.processor.get_and_clear_log(): self._log(m)

                self.processor.read_excel(xl_f)
                for m in self.processor.get_and_clear_log(): self._log(m)

                csv_folder = os.path.dirname(csv_f) if self.processor.tubos_repetidos else None
                n = self.processor.merge_data(csv_folder=csv_folder)
                for m in self.processor.get_and_clear_log(): self._log(m)

                zone = self.utm_zone.get(); epsg = 32600 + zone
                base_dir = os.path.dirname(xl_f)
                base_name = os.path.splitext(os.path.basename(xl_f))[0]

                gis_folder = os.path.join(os.path.dirname(base_dir), 'GIS') if 'PREDOBLADO' in base_dir.upper() else os.path.join(base_dir, 'GIS')
                script_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in dir() else os.getcwd()
                project_root = os.path.dirname(script_dir)
                cfg = self.processor.config
                if cfg:
                    gj_rel = cfg.get('ARCHIVOS', 'geojson_salida', fallback='')
                    if gj_rel:
                        gis_folder = os.path.dirname(os.path.join(project_root, gj_rel))

                tramo_result = self.processor.export_by_tramo(gis_folder, epsg)
                for m in self.processor.get_and_clear_log(): self._log(m)

                if n == 0 and not tramo_result:
                    self._log("\nX No se encontraron coincidencias.")
                    messagebox.showwarning("Sin resultados","No se encontraron tubos en comun.\nVerifique columna A del Excel y columna E del CSV.")
                    return

                fmt = self.output_format.get()
                self._log(f"\nExportando combinado (UTM Zona {zone}N - EPSG:{epsg})...")
                archivos = []

                if n > 0:
                    if fmt in ('geojson','all'):
                        p = os.path.join(gis_folder, f"{base_name}_GIS.geojson")
                        self.processor.export_geojson(p, epsg)
                        for m in self.processor.get_and_clear_log(): self._log(m)
                        archivos.append(p)
                    if fmt in ('kml','all'):
                        p = os.path.join(gis_folder, f"{base_name}_GIS.kml")
                        self.processor.export_kml(p, zone)
                        for m in self.processor.get_and_clear_log(): self._log(m)
                        archivos.append(p)
                    if fmt in ('csv','all'):
                        p = os.path.join(gis_folder, f"{base_name}_GIS.csv")
                        self.processor.export_csv(p)
                        for m in self.processor.get_and_clear_log(): self._log(m)
                        archivos.append(p)

                con_coords = sum(1 for v in tramo_result.values() if v['type'] == 'geojson')
                sin_coords = sum(1 for v in tramo_result.values() if v['type'] == 'csv')
                self._log("\n" + "=" * 50)
                self._log(f"LISTO - {len(tramo_result)} tramos ({con_coords} con coords, {sin_coords} sin coords)")
                if n > 0:
                    self._log(f"  Tubos combinados: {n}")
                for a in archivos: self._log(f"  -> {os.path.basename(a)}")
                self._log("=" * 50)
                self._log(f"\nArchivos por tramo en: {gis_folder}")
                self._log("Abra los .geojson en QGIS para verlos")

                if messagebox.askyesno("Listo", f"{len(tramo_result)} tramos exportados.\n{con_coords} con coordenadas, {sin_coords} sin coordenadas.\n\nArchivos en:\n{gis_folder}\n\nAbrir carpeta?"):
                    if sys.platform == 'win32': os.startfile(gis_folder)

            except Exception as ex:
                self._log(f"\nX ERROR: {ex}")
                import traceback; self._log(traceback.format_exc())
                messagebox.showerror("Error", str(ex))

    if not TIENE_OPENPYXL:
        try:
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
            import openpyxl; TIENE_OPENPYXL = True
        except: pass

    app = App()
    app.mainloop()
